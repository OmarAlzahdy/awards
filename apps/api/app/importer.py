from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete, func, select, update
from sqlalchemy.orm import Session

from .config import get_settings
from .database import build_runtime
from .models import Award, Base, Dataset, ImportIssue, Winner

ARABIC_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
ARABIC_PUNCTUATION = str.maketrans({"،": ",", "؛": ";", "ـ": " "})


@dataclass
class ImportIssuePayload:
    source_sheet: str
    source_key: str
    issue_type: str
    raw_row_json: str


@dataclass
class ImportResult:
    awards_imported: int
    winners_imported: int
    issues_recorded: int
    ignored_blank_columns: int
    report_path: str


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = text.translate(ARABIC_DIGITS).translate(ARABIC_PUNCTUATION)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_search_text(*values: Any) -> str:
    parts: list[str] = []
    for value in values:
        normalized = normalize_text(value)
        if normalized:
            parts.append(normalized.lower())
    return " ".join(parts)


def parse_int(value: Any) -> int | None:
    if isinstance(value, float) and value == int(value):
        return int(value)
    normalized = normalize_text(value)
    if not normalized:
        return None
    # Take only the first unbroken digit sequence to avoid mangling "2002.0" → 20020
    match = re.search(r"-?\d+", normalized)
    if not match:
        return None
    try:
        return int(match.group(0))
    except ValueError:
        return None


def row_to_json(headers: list[str], row: tuple[Any, ...]) -> str:
    payload = {header: normalize_text(cell) for header, cell in zip(headers, row)}
    return json.dumps(payload, ensure_ascii=False)


def _clean_header_row(header_row: tuple[Any, ...]) -> tuple[list[str], int]:
    headers: list[str] = []
    ignored_blank_columns = 0
    for cell in header_row:
        header = normalize_text(cell)
        if header:
            headers.append(header)
        else:
            ignored_blank_columns += 1
    return headers, ignored_blank_columns


def _award_to_dict(award: Award) -> dict:
    return {
        "id": award.id,
        "name": award.name,
        "summary": award.summary,
        "supervising_body": award.supervising_body,
        "prize_value": award.prize_value,
        "year_established": award.year_established,
        "country": award.country,
        "discipline": award.discipline,
        "notes": award.notes,
        "website_url": award.website_url,
        "authority_name": award.authority_name,
        "authority_type": award.authority_type,
        "search_text": award.search_text,
    }


def _winner_to_dict(winner: Winner) -> dict:
    return {
        "id": winner.id,
        "award_id": winner.award_id,
        "cycle_label": winner.cycle_label,
        "winner_name": winner.winner_name,
        "nationality_or_location": winner.nationality_or_location,
        "summary": winner.summary,
        "discipline": winner.discipline,
    }


def _create_dataset(
    session: Session,
    label: str,
    awards_data: list[dict],
    winners_data: list[dict],
    is_active: bool,
) -> Dataset:
    session.execute(update(Dataset).values(is_active=False))
    dataset = Dataset(
        label=label,
        is_active=is_active,
        imported_at=datetime.now(timezone.utc),
        awards_count=len(awards_data),
        winners_count=len(winners_data),
        awards_json=json.dumps(awards_data, ensure_ascii=False),
        winners_json=json.dumps(winners_data, ensure_ascii=False),
    )
    session.add(dataset)
    session.flush()
    return dataset


def snapshot_current_data(session: Session, label: str) -> Dataset | None:
    """Snapshot live awards+winners as a named dataset. Returns None if no awards exist."""
    awards = list(session.scalars(select(Award)))
    if not awards:
        return None
    winners = list(session.scalars(select(Winner)))
    return _create_dataset(
        session,
        label,
        [_award_to_dict(a) for a in awards],
        [_winner_to_dict(w) for w in winners],
        is_active=False,
    )


def restore_dataset(session: Session, dataset: Dataset) -> None:
    """Replace live data with the contents of the given dataset snapshot."""
    session.execute(delete(Winner))
    session.execute(delete(ImportIssue))
    session.execute(delete(Award))
    session.flush()

    awards_data: list[dict] = json.loads(dataset.awards_json)
    winners_data: list[dict] = json.loads(dataset.winners_json)

    for award_dict in awards_data:
        session.add(Award(**award_dict))
    session.flush()

    for winner_dict in winners_data:
        session.add(Winner(**winner_dict))

    session.execute(update(Dataset).values(is_active=False))
    dataset.is_active = True


def import_workbook(
    session: Session,
    workbook_path: Path,
    report_path: Path,
    label: str | None = None,
) -> ImportResult:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    awards_sheet = workbook[sheet_names[0]]
    winners_sheet = workbook[sheet_names[1]]
    websites_sheet = workbook[sheet_names[2]]
    authorities_sheet = workbook[sheet_names[3]]

    award_rows = list(awards_sheet.iter_rows(values_only=True))
    if not award_rows:
        raise ValueError("Workbook has no rows in awards sheet")

    award_headers, ignored_blank_columns = _clean_header_row(award_rows[0])
    awards_data = [
        row[: len(award_headers)]
        for row in award_rows[1:]
        if row and normalize_text(row[0])
    ]

    award_id_set = {int(row[0]) for row in awards_data if row[0] is not None}

    website_rows = list(websites_sheet.iter_rows(values_only=True))
    authority_rows = list(authorities_sheet.iter_rows(values_only=True))
    winners_rows = list(winners_sheet.iter_rows(values_only=True))

    issues: list[ImportIssuePayload] = []
    websites_by_award: dict[int, str] = {}
    authority_by_award: dict[int, tuple[str, str]] = {}

    website_headers = [normalize_text(cell) for cell in website_rows[0][:2]]
    for row in website_rows[1:]:
        if not row or not normalize_text(row[0]):
            continue
        award_id = int(row[0])
        if award_id not in award_id_set:
            issues.append(
                ImportIssuePayload(
                    source_sheet=websites_sheet.title,
                    source_key=str(award_id),
                    issue_type="orphan_website_row",
                    raw_row_json=row_to_json(website_headers, row[: len(website_headers)]),
                )
            )
            continue
        website_value = normalize_text(row[1])
        if website_value:
            websites_by_award[award_id] = website_value

    authority_headers = [normalize_text(cell) for cell in authority_rows[0][:3]]
    for row in authority_rows[1:]:
        if not row or not normalize_text(row[0]):
            continue
        award_id = int(row[0])
        if award_id not in award_id_set:
            issues.append(
                ImportIssuePayload(
                    source_sheet=authorities_sheet.title,
                    source_key=str(award_id),
                    issue_type="orphan_authority_row",
                    raw_row_json=row_to_json(authority_headers, row[: len(authority_headers)]),
                )
            )
            continue
        authority_by_award[award_id] = (normalize_text(row[1]) or "", normalize_text(row[2]) or "")

    # Snapshot existing data before wiping (only if any exists)
    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")
    existing_count = session.scalar(select(func.count(Award.id))) or 0
    if existing_count > 0:
        existing_awards = list(session.scalars(select(Award)))
        existing_winners = list(session.scalars(select(Winner)))
        prev_label = f"بيانات سابقة - {now_str}"
        _create_dataset(
            session,
            prev_label,
            [_award_to_dict(a) for a in existing_awards],
            [_winner_to_dict(w) for w in existing_winners],
            is_active=False,
        )

    session.execute(delete(Winner))
    session.execute(delete(ImportIssue))
    session.execute(delete(Award))
    session.flush()

    awards_snapshot: list[dict] = []
    awards_imported = 0
    for row in awards_data:
        award_id = int(row[0])
        authority_name, authority_type = authority_by_award.get(award_id, ("", ""))
        values = list(row) + [""] * max(0, len(award_headers) - len(row))
        search_text = normalize_search_text(
            values[1],
            values[2],
            values[3],
            values[4],
            values[6],
            values[7],
            values[8],
            authority_name,
            authority_type,
            websites_by_award.get(award_id),
        )
        award = Award(
            id=award_id,
            name=normalize_text(values[1]),
            summary=normalize_text(values[2]) or None,
            supervising_body=normalize_text(values[3]) or None,
            prize_value=normalize_text(values[4]) or None,
            year_established=parse_int(values[5]),
            country=normalize_text(values[6]) or None,
            discipline=normalize_text(values[7]) or None,
            notes=normalize_text(values[8]) or None,
            website_url=websites_by_award.get(award_id) or None,
            authority_name=authority_name or None,
            authority_type=authority_type or None,
            search_text=search_text,
        )
        session.add(award)
        awards_snapshot.append(_award_to_dict(award))
        awards_imported += 1

    winners_headers = [normalize_text(cell) for cell in winners_rows[0][:6]]
    winners_snapshot: list[dict] = []
    winners_imported = 0
    for row in winners_rows[1:]:
        if not row or not normalize_text(row[0]):
            continue
        award_id = int(row[0])
        if award_id not in award_id_set:
            issues.append(
                ImportIssuePayload(
                    source_sheet=winners_sheet.title,
                    source_key=str(award_id),
                    issue_type="orphan_winner_row",
                    raw_row_json=row_to_json(winners_headers, row[: len(winners_headers)]),
                )
            )
            continue
        padded = tuple(row) + (None,) * max(0, 6 - len(row))
        winner = Winner(
            award_id=award_id,
            cycle_label=normalize_text(padded[1]) or None,
            winner_name=normalize_text(padded[2]),
            nationality_or_location=normalize_text(padded[3]) or None,
            summary=normalize_text(padded[4]) or None,
            discipline=normalize_text(padded[5]) or None,
        )
        session.add(winner)
        winners_imported += 1

    for issue in issues:
        session.add(ImportIssue(**asdict(issue)))

    session.flush()

    # Collect winner IDs after flush so they are assigned
    flushed_winners = list(session.scalars(select(Winner)))
    for w in flushed_winners:
        winners_snapshot.append(_winner_to_dict(w))

    # Save new data as active dataset snapshot
    dataset_label = label or f"استيراد {Path(workbook_path).name} - {now_str}"
    new_dataset = Dataset(
        label=dataset_label,
        is_active=True,
        imported_at=datetime.now(timezone.utc),
        awards_count=awards_imported,
        winners_count=winners_imported,
        awards_json=json.dumps(awards_snapshot, ensure_ascii=False),
        winners_json=json.dumps(winners_snapshot, ensure_ascii=False),
    )
    session.add(new_dataset)

    session.commit()

    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_payload = {
        "workbook": str(workbook_path),
        "awards_imported": awards_imported,
        "winners_imported": winners_imported,
        "issues_recorded": len(issues),
        "ignored_blank_columns": ignored_blank_columns,
        "issues": [asdict(issue) for issue in issues],
    }
    report_path.write_text(json.dumps(report_payload, ensure_ascii=False, indent=2), encoding="utf-8")

    return ImportResult(
        awards_imported=awards_imported,
        winners_imported=winners_imported,
        issues_recorded=len(issues),
        ignored_blank_columns=ignored_blank_columns,
        report_path=str(report_path),
    )


def ensure_seed_data(session: Session, workbook_path: Path, report_path: Path) -> None:
    existing_award = session.scalar(select(Award.id).limit(1))
    if existing_award is None and workbook_path.exists():
        import_workbook(session, workbook_path, report_path, label="البيانات الأصلية")
    else:
        # If data exists but no dataset snapshots yet, create an initial snapshot
        existing_datasets = session.scalar(select(func.count(Dataset.id))) or 0
        if existing_datasets == 0:
            awards = list(session.scalars(select(Award)))
            if awards:
                winners = list(session.scalars(select(Winner)))
                _create_dataset(
                    session,
                    "البيانات الأصلية",
                    [_award_to_dict(a) for a in awards],
                    [_winner_to_dict(w) for w in winners],
                    is_active=True,
                )
                session.commit()


def main() -> None:
    settings = get_settings()
    runtime = build_runtime(settings)
    Base.metadata.create_all(runtime.engine)
    with runtime.session_factory() as session:
        result = import_workbook(session, settings.workbook_file, settings.report_file)
    print(json.dumps(asdict(result), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
